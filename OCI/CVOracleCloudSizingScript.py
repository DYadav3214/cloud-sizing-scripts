import sys
import subprocess
import oci
from oci.monitoring.models import SummarizeMetricsDataDetails
import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
import os
from datetime import datetime
import logging
import tempfile
import json
import shutil

oci_path = shutil.which("oci")
kubectl_path = shutil.which("kubectl")

total_instances = 0
total_instance_sizeGB = 0
total_instance_sizeTB = 0
total_namespaces = 0
total_buckets = 0
total_storageGB = 0
total_storageTB = 0
total_db_systems = 0
total_db_system_sizeGB = 0
total_db_system_sizeTB = 0
total_oke_clusters = 0
total_oke_node_count = 0
total_oke_pvc_count = 0
total_oke_pvc_size_gb = 0
total_oke_pvc_size_tb = 0
total_oke_statefulsets = 0
total_oke_unprotected = 0

class InstanceInfo:
    def __init__(self):
        self.compartment_id = None
        self.instance_id = None
        self.instance_name = None
        self.region = None
        self.availability_domain = None
        self.shape = None
        self.shape_family = None         # NEW: VM.Standard | BM.Standard | VM.GPU etc.
        self.ocpus = 0                   # NEW: CPU count
        self.memory_gb = 0               # NEW: RAM GB
        self.state = None
        self.number_of_volumes = 0
        self.sizeGB = 0
        self.sizeTB = 0
        self.defined_tags = {}
        self.freeform_tags = {}
        self.boot_volume_name = None
        self.block_volume_names = []
        self.os_type = None              # NEW: Oracle Linux | Windows | Ubuntu etc.
        self.backup_policy = None        # NEW: Gold | Silver | Bronze | None
        self.has_backup_policy = False   # NEW
        self.is_windows = False          # NEW: drives agent licensing estimation

class InstanceSummary:
    def __init__(self):
        self.region = None
        self.instance_count = 0
        self.total_sizeGB = 0
        self.total_sizeTB = 0

class ObjectStorageInfo:
    def __init__(self):
        self.compartment_id = None
        self.namespace = None
        self.bucket_name = None
        self.region = None
        self.storage_tier = None
        self.object_count = 0
        self.sizeGB = 0
        self.sizeTB = 0
        self.defined_tags = {}
        self.freeform_tags = {}
        self.versioning_enabled = False  # NEW: versioning = some protection, but not backup
        self.replication_enabled = False # NEW: cross-region replication configured
        self.lifecycle_rules_count = 0   # NEW: auto-delete rules = data at risk
        self.public_access = False       # NEW: data exposure risk

class ObjectStorageSummary:
    def __init__(self):
        self.region = None
        self.namespace = None
        self.bucket_count = 0
        self.total_storage_GB = 0
        self.total_storage_TB = 0

class DBSystemInfo:
    def __init__(self):
        self.compartment_id = None
        self.db_system_id = None
        self.display_name = None
        self.region = None
        self.availability_domain = None
        self.shape = None
        self.lifecycle_state = None
        self.node_count = 0
        self.db_version = None
        self.database_edition = None
        self.data_storage_size_gb = 0
        self.data_storage_size_tb = 0
        self.defined_tags = {}
        self.freeform_tags = {}
        self.backup_config = None        # NEW: AUTOBACKUP_ENABLED etc.
        self.reco_storage_size_gb = 0    # NEW: recovery storage allocated
        self.license_model = None        # NEW: LICENSE_INCLUDED | BRING_YOUR_OWN_LICENSE

class DBSystemSummary:
    def __init__(self):
        self.region = None
        self.db_system_count = 0
        self.total_storage_gb = 0
        self.total_storage_tb = 0

class OKEClusterInfo:
    def __init__(self):
        self.region = None
        self.cluster_id = None
        self.cluster_name = None
        self.kubernetes_version = None
        self.lifecycle_state = None
        # Nodes
        self.node_count = 0
        self.node_names = []
        self.node_shapes = []            # NEW: instance shapes (e.g. VM.Standard.E4.Flex)
        self.node_os_images = []         # NEW: OS image per node
        # PVCs
        self.pvc_count = 0
        self.pvc_names = []
        self.total_pvc_size_gb = 0
        self.total_pvc_size_tb = 0
        # NEW: enriched PVC/storage signals
        self.storage_classes = []        # unique StorageClasses in use
        self.pvc_by_storage_class = {}   # {storageClass: {count, size_gb}}
        self.pvc_unbound_count = 0       # PVCs not bound (data at risk)
        # NEW: workload signals
        self.namespace_count = 0
        self.namespaces = []
        self.statefulset_count = 0       # StatefulSets = most critical backup targets
        self.statefulset_names = []
        self.deployment_count = 0
        self.has_backup_annotations = False   # any velero/trilio/kasten annotation found
        self.backup_tool_detected = None      # velero | trilio | kasten | none
        self.configmap_count = 0
        self.secret_count = 0            # proxy for app config complexity
        # NEW: data classification signals (from namespace names / labels)
        self.likely_production = False
        self.likely_database_workloads = []  # namespaces containing db-like apps
        # NEW: node pool info
        self.node_pool_count = 0
        self.node_pool_shapes = []

class OKEClusterSummary:
    def __init__(self):
        self.region = None
        self.cluster_count = 0
        self.total_node_count = 0
        self.total_pvc_count = 0
        self.total_pvc_size_gb = 0
        self.total_pvc_size_tb = 0
        self.total_statefulsets = 0
        self.unprotected_clusters = 0   # no backup tool detected


def install_and_import(package):
    try:
        __import__(package)
        print(f"Package '{package}' is already installed.")
    except ImportError:
        print(f"Package '{package}' not found. Installing...")
        subprocess.check_call([sys.executable, "-m", "pip", "install", package])

def get_sheet_info(workload):
    if workload == "instances":
        info_sheet = "Instance Info"
        summary_sheet = "Instance Summary"
        info_headers = [
            "Compartment ID", "Instance ID", "Instance Name", "Region",
            "Availability Domain", "Shape", "Shape Family", "OCPUs", "Memory (GB)",
            "State", "OS Type", "Is Windows",
            "Number of Volumes", "Size (GB)", "Size (TB)",
            "Boot Volume Name", "Block Volume Names",
            "Has Backup Policy", "Backup Policy",
            "Defined Tags", "Freeform Tags",
        ]
        summary_headers = ["Region", "Instance Count", "Total Size (GB)", "Total Size (TB)"]
    elif workload == "object_storage":
        info_sheet = "Object Storage Info"
        summary_sheet = "Object Storage Summary"
        info_headers = [
            "Namespace", "Compartment ID", "Bucket Name", "Region",
            "Storage Tier", "Object Count", "Size (GB)", "Size (TB)",
            "Versioning Enabled", "Replication Enabled", "Lifecycle Rules", "Public Access",
            "Defined Tags", "Freeform Tags",
        ]
        summary_headers = ["Region", "Bucket Count", "Total Size (GB)", "Total Size (TB)"]
    elif workload == "db_systems":
        info_sheet = "DB System Info"
        summary_sheet = "DB System Summary"
        info_headers = [
            "Compartment ID", "DB System ID", "Display Name", "Region",
            "Availability Domain", "Shape", "Lifecycle State", "Node Count",
            "DB Version", "Database Edition", "Data Storage Size (GB)", "Data Storage Size (TB)",
            "Recovery Storage (GB)", "License Model", "Backup Config",
            "Defined Tags", "Freeform Tags",
        ]
        summary_headers = ["Region", "DB System Count", "Total Storage (GB)", "Total Storage (TB)"]
    elif workload == "oke_clusters":
        info_sheet = "OKE Cluster Info"
        summary_sheet = "OKE Cluster Summary"
        info_headers = [
            "Region", "Cluster ID", "Cluster Name", "Kubernetes Version", "Lifecycle State",
            # Nodes
            "Node Count", "Node Shapes",
            # PVCs
            "PVC Count", "Total PVC Size (GB)", "Total PVC Size (TB)",
            "Storage Classes", "Unbound PVCs",
            "PVC by StorageClass",
            # Workloads
            "Namespaces", "Namespace Count",
            "StatefulSet Count", "StatefulSets",
            "Deployment Count",
            "ConfigMap Count", "Secret Count",
            # Protection
            "Backup Tool Detected", "Has Backup Annotations",
            "Likely Production", "Likely DB Workload Namespaces",
            # Raw lists
            "PVC Names", "Node Names",
        ]
        summary_headers = [
            "Region", "Cluster Count", "Total Node Count",
            "Total PVC Count", "Total PVC Size (GB)", "Total PVC Size (TB)",
            "Total StatefulSets", "Unprotected Clusters",
        ]
    else:
        raise ValueError(f"Unsupported workload: {workload}")

    return info_sheet, summary_sheet, info_headers, summary_headers

def format_workbook(filename):
    wb = load_workbook(filename)
    header_fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
    bold_font = Font(bold=True)

    for sheet in wb.worksheets:
        # Format header row
        for cell in sheet[1]:
            cell.font = bold_font
            cell.fill = header_fill

        # Adjust column widths
        for column_cells in sheet.columns:
            max_length = 0
            col = column_cells[0].column_letter
            for cell in column_cells:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            adjusted_width = max_length + 2
            sheet.column_dimensions[col].width = adjusted_width
    wb.save(filename)

def init_excel(filename, workload):
    if not os.path.exists(filename):
        wb = Workbook()
        wb.remove(wb.active)  # Remove default sheet
        
        info_sheet, summary_sheet, info_headers, summary_headers = get_sheet_info(workload)

        wb.create_sheet(summary_sheet)
        wb.create_sheet(info_sheet)

        wb[info_sheet].append(info_headers)
        wb[summary_sheet].append(summary_headers)
        wb.save(filename)

def dump_info(filename, workload, object_list: list, include_sensitive: bool = False):
    info_sheet, _, _, _ = get_sheet_info(workload)
    wb = load_workbook(filename)
    sheet = wb[info_sheet]

    for obj in object_list:
        if workload == "instances":
            row = [
                obj.compartment_id,
                obj.instance_id,
                obj.instance_name,
                obj.region,
                obj.availability_domain,
                obj.shape,
                obj.shape_family or "",
                obj.ocpus,
                obj.memory_gb,
                obj.state,
                obj.os_type or "",
                str(obj.is_windows),
                obj.number_of_volumes,
                obj.sizeGB,
                obj.sizeTB,
                obj.boot_volume_name if obj.boot_volume_name else "",
                ", ".join(obj.block_volume_names) if obj.block_volume_names else "",
                str(obj.has_backup_policy),
                obj.backup_policy or "none",
                str(obj.defined_tags),
                str(obj.freeform_tags),
            ]
        elif workload == "object_storage":
            row = [
                obj.namespace,
                obj.compartment_id,
                obj.bucket_name,
                obj.region,
                obj.storage_tier,
                obj.object_count,
                obj.sizeGB,
                obj.sizeTB,
                str(obj.versioning_enabled),
                str(obj.replication_enabled),
                obj.lifecycle_rules_count,
                str(obj.public_access),
                str(obj.defined_tags),
                str(obj.freeform_tags),
            ]
        elif workload == "db_systems":
            row = [
                obj.compartment_id,
                obj.db_system_id,
                obj.display_name,
                obj.region,
                obj.availability_domain,
                obj.shape,
                obj.lifecycle_state,
                obj.node_count,
                obj.db_version,
                obj.database_edition,
                obj.data_storage_size_gb,
                obj.data_storage_size_tb,
                obj.reco_storage_size_gb,
                obj.license_model or "",
                obj.backup_config or "unknown",
                str(obj.defined_tags),
                str(obj.freeform_tags),
            ]
        elif workload == "oke_clusters":
            row = [
                obj.region,
                obj.cluster_id,
                obj.cluster_name,
                obj.kubernetes_version,
                obj.lifecycle_state or "",
                obj.node_count,
                ", ".join(sorted(set(obj.node_shapes))) if obj.node_shapes else "",
                obj.pvc_count,
                obj.total_pvc_size_gb,
                obj.total_pvc_size_tb,
                ", ".join(sorted(set(obj.storage_classes))) if obj.storage_classes else "",
                obj.pvc_unbound_count,
                str(obj.pvc_by_storage_class) if obj.pvc_by_storage_class else "",
                # Sensitive: namespace names, StatefulSet names, PVC names, node names
                ", ".join(obj.namespaces) if (include_sensitive and obj.namespaces) else "[redacted]",
                obj.namespace_count,
                obj.statefulset_count,
                ", ".join(obj.statefulset_names) if (include_sensitive and obj.statefulset_names) else "[redacted]",
                obj.deployment_count,
                obj.configmap_count,
                obj.secret_count,
                obj.backup_tool_detected or "none",
                str(obj.has_backup_annotations),
                str(obj.likely_production),
                ", ".join(obj.likely_database_workloads) if obj.likely_database_workloads else "",
                ", ".join(obj.pvc_names[:50]) if (include_sensitive and obj.pvc_names) else "[redacted]",
                ", ".join(obj.node_names) if (include_sensitive and obj.node_names) else "[redacted]",
            ]
        else:
            raise ValueError(f"Unsupported workload: {workload}")
        sheet.append(row)
    wb.save(filename)

def dump_summary(filename, workload, summary):
    _, summary_sheet, _, _ = get_sheet_info(workload)
    wb = load_workbook(filename)
    sheet = wb[summary_sheet]

    for obj in summary:
        if workload == "instances":
            row = [
                obj.region,
                obj.instance_count,
                obj.total_sizeGB,
                obj.total_sizeTB,
            ]
        elif workload == "object_storage":
            row = [
                obj.region,
                obj.bucket_count,
                obj.total_storage_GB,
                obj.total_storage_TB,
            ]
        elif workload == "db_systems":
            row = [
                obj.region,
                obj.db_system_count,
                obj.total_storage_gb,
                obj.total_storage_tb,
            ]
        elif workload == "oke_clusters":
            row = [
                obj.region,
                obj.cluster_count,
                obj.total_node_count,
                obj.total_pvc_count,
                obj.total_pvc_size_gb,
                obj.total_pvc_size_tb,
                obj.total_statefulsets,
                obj.unprotected_clusters,
            ]
        else:
            raise ValueError(f"Unsupported workload: {workload}")
        sheet.append(row)

    wb.save(filename)

def write_grand_total(filename, workload):
    wb = load_workbook(filename)

    bold_font = Font(bold=True)
    if workload == "instances":
        global total_instances, total_instance_sizeGB, total_instance_sizeTB
        sheet_name = "Instance Summary"
        row = ["Total Instances", total_instances, total_instance_sizeGB, total_instance_sizeTB]
    elif workload == "object_storage":
        sheet_name = "Object Storage Summary"
        global total_namespaces, total_buckets, total_storageGB, total_storageTB
        row = ["Total Buckets", total_buckets, total_storageGB, total_storageTB]
    elif workload == "db_systems":
        sheet_name = "DB System Summary"
        global total_db_systems, total_db_system_sizeGB, total_db_system_sizeTB
        row = ["Total DB Systems", total_db_systems, total_db_system_sizeGB, total_db_system_sizeTB]
    elif workload == "oke_clusters":
        sheet_name = "OKE Cluster Summary"
        global total_oke_clusters, total_oke_node_count, total_oke_pvc_count
        global total_oke_pvc_size_gb, total_oke_pvc_size_tb
        global total_oke_statefulsets, total_oke_unprotected
        row = [
            "Total OKE Clusters",
            total_oke_clusters,
            total_oke_node_count,
            total_oke_pvc_count,
            total_oke_pvc_size_gb,
            total_oke_pvc_size_tb,
            total_oke_statefulsets,
            total_oke_unprotected,
        ]
    else:
        raise ValueError(f"Unsupported workload: {workload}")

    sheet = wb[sheet_name]
    sheet.append(row)
    for cell in sheet[2]:
        cell.font = bold_font
    wb.save(filename)

def get_object_storage_info(config, filename, regions=[], compartments=[]):
    global total_buckets, total_storageGB, total_storageTB
    object_storage_summary_list = []
    identity_client = oci.identity.IdentityClient(config)
    if not regions:
        regions = [region.region_name for region in identity_client.list_region_subscriptions(config["tenancy"]).data]
    if not compartments:
        compartments = [compartment.id for compartment in identity_client.list_compartments(compartment_id=config["tenancy"], compartment_id_in_subtree=True).data]
    for region in regions:
        logging.info(f"Processing region: {region}")
        config["region"] = region
        object_storage_client = oci.object_storage.ObjectStorageClient(config)
        region_summary = ObjectStorageSummary()
        region_summary.region = region
        region_summary.bucket_count = 0
        region_summary.total_storage_GB = 0
        region_summary.total_storage_TB = 0
        try:
            namespace = object_storage_client.get_namespace().data
        except Exception as e:
            logging.error(f"Error fetching namespace for region {region}: {e}")
            continue
        for compartment in compartments:
            compartment_bucket_list = []
            try:
                buckets = oci.pagination.list_call_get_all_results(
                    object_storage_client.list_buckets,
                    namespace_name=namespace,
                    compartment_id=compartment
                ).data
            except Exception as e:
                logging.error(f"Error fetching buckets for compartment {compartment}: {e}")
                continue
            logging.info(f"Found {len(buckets)} bucket(s) in compartment {compartment}")
            if len(buckets) == 0:
                continue
            for bucket in buckets:
                bucket_info = ObjectStorageInfo()
                bucket_info.compartment_id = compartment
                bucket_info.namespace = namespace
                bucket_info.bucket_name = bucket.name
                bucket_info.region = region
                try:
                    stats = object_storage_client.get_bucket(
                        namespace_name=namespace,
                        bucket_name=bucket.name,
                        fields=['approximateSize', 'approximateCount']
                    ).data
                    bucket_info.storage_tier = stats.storage_tier
                    bucket_info.defined_tags = stats.defined_tags
                    bucket_info.freeform_tags = stats.freeform_tags
                    size_in_bytes, object_count = stats.approximate_size, stats.approximate_count
                    bucket_info.sizeGB = round(size_in_bytes / (1024 ** 3), 2) if size_in_bytes else 0
                    bucket_info.sizeTB = round(bucket_info.sizeGB / 1024, 2) if bucket_info.sizeGB else 0
                    bucket_info.object_count = object_count
                    # Versioning
                    bucket_info.versioning_enabled = getattr(stats, "versioning", None) == "Enabled"
                    # Public access
                    bucket_info.public_access = getattr(stats, "public_access_type", "NoPublicAccess") != "NoPublicAccess"
                except Exception as e:
                    logging.error(f"Error fetching stats for bucket {bucket.name}: {e}")
                    continue
                # Replication (best-effort)
                try:
                    replication = oci.pagination.list_call_get_all_results(
                        object_storage_client.list_replication_policies,
                        namespace_name=namespace, bucket_name=bucket.name,
                    ).data
                    bucket_info.replication_enabled = len(replication) > 0
                except Exception:
                    bucket_info.replication_enabled = False
                # Lifecycle rules
                try:
                    lc = object_storage_client.get_object_lifecycle_policy(
                        namespace_name=namespace, bucket_name=bucket.name,
                    ).data
                    bucket_info.lifecycle_rules_count = len(lc.items) if lc and lc.items else 0
                except Exception:
                    bucket_info.lifecycle_rules_count = 0
                region_summary.bucket_count += 1
                region_summary.total_storage_GB += bucket_info.sizeGB if bucket_info.sizeGB else 0
                region_summary.total_storage_TB += bucket_info.sizeTB if bucket_info.sizeTB else 0
                compartment_bucket_list.append(bucket_info)
                total_buckets += 1
                total_storageGB += bucket_info.sizeGB if bucket_info.sizeGB else 0
                total_storageTB += bucket_info.sizeTB if bucket_info.sizeTB else 0
            dump_info(filename, "object_storage", compartment_bucket_list, include_sensitive)
        object_storage_summary_list.append(region_summary)
        
    write_grand_total(filename, "object_storage")
    dump_summary(filename, "object_storage", object_storage_summary_list)
    format_workbook(filename)
    logging.info("Completed processing all regions and compartments for object storage.")
    logging.info(f"Grand Total - Buckets: {total_buckets}, Size (GB): {total_storageGB}, Size (TB): {total_storageTB}")

def get_boot_volume_info(config, instance_id, availability_domain, compartment_id):
    compute_client = oci.core.ComputeClient(config)
    block_storage_client = oci.core.BlockstorageClient(config)
    try:
        response = oci.pagination.list_call_get_all_results(compute_client.list_boot_volume_attachments,
                                                            instance_id=instance_id,
                                                            availability_domain=availability_domain,
                                                            compartment_id=compartment_id,
                                                            retry_strategy=oci.retry.DEFAULT_RETRY_STRATEGY)
        boot_volumes = response.data
        result = {"name": None, "sizeGB": 0}
        if not boot_volumes:
            return result
        try:
            response = block_storage_client.get_boot_volume(boot_volumes[0].boot_volume_id)
            boot_volume_info = response.data
            result = {"name": boot_volume_info.display_name, "sizeGB": boot_volume_info.size_in_gbs}
            return result
        except Exception as e:
            print(f"Error retrieving boot volume info for instance {instance_id}: {e}")
            return result
    except Exception as e:
        print(f"Error retrieving boot volume attachments for instance {instance_id}: {e}")
        return {"name": None, "sizeGB": 0}

def get_block_volume_info(config, instance_id, availability_domain, compartment_id):
    compute_client = oci.core.ComputeClient(config)
    block_storage_client = oci.core.BlockstorageClient(config)
    try:
        response = oci.pagination.list_call_get_all_results(compute_client.list_volume_attachments,
                                                            instance_id=instance_id,
                                                            availability_domain=availability_domain,
                                                            compartment_id=compartment_id,
                                                            retry_strategy=oci.retry.DEFAULT_RETRY_STRATEGY)
        volume_attachments = response.data
        result = []
        if not volume_attachments:
            return result
        for attachment in volume_attachments:
            try:
                response = block_storage_client.get_volume(attachment.volume_id)
                volume_info = response.data
                result.append({"name": volume_info.display_name, "sizeGB": volume_info.size_in_gbs})
            except Exception as e:
                print(f"Error retrieving volume info for volume {attachment.volume_id}: {e}")
        return result
    except Exception as e:
        print(f"Error retrieving volume attachments for instance {instance_id}: {e}")
        return []

def get_instance_info(config, filename, regions=[], compartments=[]):
    global total_instances, total_instance_sizeGB, total_instance_sizeTB
    instance_summary_list = []
    identity_client = oci.identity.IdentityClient(config)
    if not regions: 
        regions = [region.region_name for region in identity_client.list_region_subscriptions(config["tenancy"]).data]
    if not compartments:
        compartments = [compartment.id for compartment in identity_client.list_compartments(compartment_id=config["tenancy"], compartment_id_in_subtree=True).data]
    for region in regions:
        logging.info(f"Processing region: {region}")
        config["region"] = region
        compute_client = oci.core.ComputeClient(config)
        region_summary = InstanceSummary()
        for compartment in compartments:
            logging.info(f"Processing compartment: {compartment}")
            compartment_instance_list = []
            region_summary.region = region
            instances = oci.pagination.list_call_get_all_results(compute_client.list_instances,
                                                                compartment_id=compartment,
                                                                retry_strategy=oci.retry.DEFAULT_RETRY_STRATEGY).data
            logging.info(f"Found {len(instances)} instance(s)")
            if len(instances) == 0:
                continue
            for instance in instances:
                if instance.lifecycle_state == "TERMINATED":
                    continue
                logging.info(f"Processing instance: {instance.id} - {instance.display_name}")
                instance_info = InstanceInfo()
                instance_info.compartment_id = compartment
                instance_info.instance_id = instance.id
                instance_info.instance_name = instance.display_name
                instance_info.region = region
                instance_info.availability_domain = instance.availability_domain
                instance_info.shape = instance.shape
                instance_info.shape_family = instance.shape.split(".")[0] if instance.shape else None
                instance_info.state = instance.lifecycle_state
                instance_info.defined_tags = instance.defined_tags
                instance_info.freeform_tags = instance.freeform_tags
                # Shape config (CPU/RAM)
                if hasattr(instance, "shape_config") and instance.shape_config:
                    instance_info.ocpus = getattr(instance.shape_config, "ocpus", 0) or 0
                    instance_info.memory_gb = getattr(instance.shape_config, "memory_in_gbs", 0) or 0
                # OS type from image (best-effort via tags/display name)
                img_name = str(instance.freeform_tags.get("image", "") or
                               instance.defined_tags.get("Oracle-Tags", {}).get("CreatedBy", "") or
                               instance.display_name or "").lower()
                if "windows" in img_name:
                    instance_info.os_type = "Windows"
                    instance_info.is_windows = True
                elif "oracle" in img_name:
                    instance_info.os_type = "Oracle Linux"
                elif "ubuntu" in img_name:
                    instance_info.os_type = "Ubuntu"
                elif "centos" in img_name:
                    instance_info.os_type = "CentOS"
                else:
                    instance_info.os_type = "Linux/Unknown"
                # Backup policy from volume backup policies
                try:
                    vol_client = oci.core.BlockstorageClient(config)
                    boot_policies = oci.pagination.list_call_get_all_results(
                        vol_client.get_volume_backup_policy_asset_assignment,
                        asset_id=instance.id
                    ).data
                    if boot_policies:
                        instance_info.backup_policy = getattr(boot_policies[0], "policy_id", "assigned")
                        instance_info.has_backup_policy = True
                except Exception:
                    instance_info.backup_policy = None
                    instance_info.has_backup_policy = False
                try:
                    boot_volume_info = get_boot_volume_info(config, instance.id, instance.availability_domain, instance.compartment_id)
                    block_volumes_info = get_block_volume_info(config, instance.id, instance.availability_domain, instance.compartment_id)
                except Exception as e:
                    logging.error(f"Error fetching volume data for instance {instance.id}: {e}")
                    continue
                instance_info.number_of_volumes = (1 if boot_volume_info["sizeGB"] > 0 else 0) + len(block_volumes_info)
                instance_info.sizeGB = boot_volume_info["sizeGB"] + sum([bv["sizeGB"] for bv in block_volumes_info])
                instance_info.sizeTB = round(instance_info.sizeGB / 1024, 2)
                instance_info.boot_volume_name = boot_volume_info["name"] if boot_volume_info["name"] else None
                instance_info.block_volume_names = [bv["name"] for bv in block_volumes_info]
                region_summary.instance_count += 1
                region_summary.total_sizeGB += instance_info.sizeGB
                region_summary.total_sizeTB += instance_info.sizeTB
                compartment_instance_list.append(instance_info)
                total_instances += 1
                total_instance_sizeGB += instance_info.sizeGB
                total_instance_sizeTB += instance_info.sizeTB
            dump_info(filename, "instances", compartment_instance_list, include_sensitive)
        instance_summary_list.append(region_summary)
    write_grand_total(filename, "instances")
    dump_summary(filename, "instances", instance_summary_list)
    format_workbook(filename)
    logging.info("Completed processing all regions and compartments.")
    logging.info(f"Grand Total - Instances: {total_instances}, Size (GB): {total_instance_sizeGB}, Size (TB): {total_instance_sizeTB}")

def get_database_info(config, filename, regions=[], compartments=[]):
    global total_db_systems, total_db_system_sizeGB, total_db_system_sizeTB
    db_summary_list = []
    identity_client = oci.identity.IdentityClient(config)
    if not regions:
        regions = [region.region_name for region in identity_client.list_region_subscriptions(config["tenancy"]).data]
    if not compartments:
        compartments = [compartment.id for compartment in identity_client.list_compartments(compartment_id=config["tenancy"], compartment_id_in_subtree=True).data]
    for region in regions:
        logging.info(f"Processing region: {region}")
        config["region"] = region
        db_client = oci.database.DatabaseClient(config)
        region_summary = DBSystemSummary()
        region_summary.region = region
        for compartment in compartments:
            logging.info(f"Processing compartment: {compartment}")
            compartment_db_list = []
            try:
                db_systems = oci.pagination.list_call_get_all_results(
                    db_client.list_db_systems,
                    compartment_id=compartment
                ).data
            except Exception as e:
                logging.error(f"Error fetching DB systems for compartment {compartment}: {e}")
                continue
            logging.info(f"Found {len(db_systems)} DB system(s)")
            if len(db_systems) == 0:
                continue
            for db in db_systems:
                if db.lifecycle_state == "TERMINATED":
                    continue
                db_info = DBSystemInfo()
                db_info.compartment_id = compartment
                db_info.db_system_id = db.id
                db_info.display_name = db.display_name
                db_info.region = region
                db_info.availability_domain = db.availability_domain
                db_info.shape = db.shape
                db_info.lifecycle_state = db.lifecycle_state
                db_info.node_count = db.node_count if hasattr(db, "node_count") else 0
                db_info.db_version = db.version if hasattr(db, "version") else ""
                db_info.database_edition = db.database_edition if hasattr(db, "database_edition") else "" 
                db_info.data_storage_size_gb = db.data_storage_size_in_gbs if hasattr(db, "data_storage_size_in_gbs") else 0
                db_info.data_storage_size_tb = round(db_info.data_storage_size_gb / 1024, 2)
                db_info.defined_tags = db.defined_tags
                db_info.freeform_tags = db.freeform_tags
                db_info.license_model = db.license_model if hasattr(db, "license_model") else None
                db_info.reco_storage_size_gb = db.reco_storage_size_in_gb if hasattr(db, "reco_storage_size_in_gb") else 0
                # Backup config (best-effort)
                try:
                    dbs_list = oci.pagination.list_call_get_all_results(
                        db_client.list_db_homes, compartment_id=compartment, db_system_id=db.id
                    ).data
                    if dbs_list:
                        dbs = oci.pagination.list_call_get_all_results(
                            db_client.list_databases, compartment_id=compartment, db_home_id=dbs_list[0].id
                        ).data
                        if dbs and hasattr(dbs[0], "db_backup_config"):
                            bc = dbs[0].db_backup_config
                            db_info.backup_config = (
                                f"auto_backup={'ENABLED' if getattr(bc, 'auto_backup_enabled', False) else 'DISABLED'}"
                                f", retention={getattr(bc, 'recovery_window_in_days', 'N/A')}d"
                            )
                        else:
                            db_info.backup_config = "unknown"
                except Exception:
                    db_info.backup_config = "unknown"
                region_summary.db_system_count += 1
                region_summary.total_storage_gb += db_info.data_storage_size_gb
                region_summary.total_storage_tb += db_info.data_storage_size_tb
                compartment_db_list.append(db_info)
                total_db_systems += 1
                total_db_system_sizeGB += db_info.data_storage_size_gb
                total_db_system_sizeTB += db_info.data_storage_size_tb
            dump_info(filename, "db_systems", compartment_db_list, include_sensitive)
        db_summary_list.append(region_summary)
    write_grand_total(filename, "db_systems")
    dump_summary(filename, "db_systems", db_summary_list)
    format_workbook(filename)
    logging.info("Completed processing all regions and compartments for DB systems.")
    logging.info(f"Grand Total - DB Systems: {total_db_systems}, Storage (GB): {total_db_system_sizeGB}, Storage (TB): {total_db_system_sizeTB}")

_BACKUP_TOOL_LABELS = {
    "velero": ["velero.io", "backup.velero.io"],
    "kasten": ["k10.kasten.io", "kas.kasten.io"],
    "trilio": ["trilio.io", "triliovault.trilio.io"],
}
_DB_NAMESPACE_HINTS = ["mysql", "postgres", "mongo", "redis", "elastic", "kafka", "cassandra",
                       "oracle", "mssql", "mariadb", "influx", "couch", "neo4j", "etcd"]
_PROD_NAMESPACE_HINTS = ["prod", "production", "prd", "live"]


def _parse_storage_gb(storage_str: str) -> float:
    """Convert Kubernetes storage string (Gi, Mi, G, M, Ti) to GB."""
    if not storage_str:
        return 0.0
    s = storage_str.strip()
    try:
        if s.endswith("Ti"):
            return float(s[:-2]) * 1024
        if s.endswith("Gi"):
            return float(s[:-2])
        if s.endswith("Mi"):
            return float(s[:-2]) / 1024
        if s.endswith("G"):
            return float(s[:-1]) * 0.931
        if s.endswith("M"):
            return float(s[:-1]) * 0.000931
        return float(s) / (1024 ** 3)
    except ValueError:
        return 0.0


def _kubectl(kubeconfig_file: str, *args) -> dict:
    """Run kubectl and return parsed JSON. Raises on error."""
    result = subprocess.run(
        ["kubectl", "--kubeconfig", kubeconfig_file] + list(args) + ["-o", "json"],
        check=True, capture_output=True, timeout=120,
    )
    # Force UTF-8 decode — kubectl output may contain non-ASCII chars in names/labels
    stdout = result.stdout.decode("utf-8", errors="replace")
    return json.loads(stdout)


def get_oke_cluster_info(config, filename, regions=[], compartments=[]):
    global total_oke_clusters, total_oke_node_count, total_oke_pvc_count
    global total_oke_pvc_size_gb, total_oke_pvc_size_tb
    global total_oke_statefulsets, total_oke_unprotected

    oke_summary_list = []
    identity_client = oci.identity.IdentityClient(config)

    if not regions:
        regions = [r.region_name for r in identity_client.list_region_subscriptions(config["tenancy"]).data]
    if not compartments:
        compartments = [c.id for c in identity_client.list_compartments(
            compartment_id=config["tenancy"], compartment_id_in_subtree=True).data]

    for region in regions:
        logging.info(f"Processing OKEs in region: {region}")
        config["region"] = region
        container_engine_client = oci.container_engine.ContainerEngineClient(config)

        region_summary = OKEClusterSummary()
        region_summary.region = region

        for compartment in compartments:
            try:
                clusters = oci.pagination.list_call_get_all_results(
                    container_engine_client.list_clusters,
                    compartment_id=compartment
                ).data
            except Exception as e:
                logging.error(f"Error fetching clusters in {compartment}: {e}")
                continue

            compartment_oke_list = []
            for cluster in clusters:
                if cluster.lifecycle_state == "DELETED":
                    continue

                cluster_info = OKEClusterInfo()
                cluster_info.region = region
                cluster_info.cluster_id = cluster.id
                cluster_info.cluster_name = cluster.name
                cluster_info.kubernetes_version = cluster.kubernetes_version
                cluster_info.lifecycle_state = cluster.lifecycle_state

                # ── OCI Node Pool info (no kubectl needed) ──────────────────
                try:
                    node_pools = oci.pagination.list_call_get_all_results(
                        container_engine_client.list_node_pools,
                        compartment_id=compartment,
                        cluster_id=cluster.id,
                    ).data
                    cluster_info.node_pool_count = len(node_pools)
                    cluster_info.node_pool_shapes = list({np.node_shape for np in node_pools if np.node_shape})
                except Exception as e:
                    logging.warning(f"Could not fetch node pools for {cluster.name}: {e}")

                kubeconfig_file = os.path.join(tempfile.gettempdir(), f"kubeconfig_{cluster.id}")
                try:
                    subprocess.run(
                        [
                            "oci", "ce", "cluster", "create-kubeconfig",
                            "--cluster-id", cluster.id,
                            "--file", kubeconfig_file,
                            "--region", region,
                            "--token-version", "2.0.0",
                            "--kube-endpoint", "PRIVATE_ENDPOINT",
                            "--profile", config.get("profile", oci.config.DEFAULT_PROFILE),
                        ],
                        check=True, capture_output=True, text=True,
                    )
                    logging.info(f"Kubeconfig created at {kubeconfig_file} for cluster {cluster.name}")

                    # ── Nodes ───────────────────────────────────────────────
                    try:
                        node_data = _kubectl(kubeconfig_file, "get", "nodes")
                        cluster_info.node_names = [n["metadata"]["name"] for n in node_data["items"]]
                        cluster_info.node_count = len(cluster_info.node_names)
                        cluster_info.node_shapes = [
                            n["metadata"].get("labels", {}).get("node.kubernetes.io/instance-type", "")
                            for n in node_data["items"]
                        ]
                        cluster_info.node_os_images = list({
                            n["status"].get("nodeInfo", {}).get("osImage", "")
                            for n in node_data["items"] if n.get("status")
                        })
                    except Exception as e:
                        logging.warning(f"Could not fetch nodes for {cluster.name}: {e}")

                    # ── Namespaces ──────────────────────────────────────────
                    try:
                        ns_data = _kubectl(kubeconfig_file, "get", "namespaces")
                        all_ns = [n["metadata"]["name"] for n in ns_data["items"]]
                        # Filter out system namespaces for signal quality
                        system_ns = {"kube-system", "kube-public", "kube-node-lease"}
                        cluster_info.namespaces = [n for n in all_ns if n not in system_ns]
                        cluster_info.namespace_count = len(cluster_info.namespaces)
                        # Data classification
                        cluster_info.likely_production = any(
                            hint in ns.lower() for ns in cluster_info.namespaces for hint in _PROD_NAMESPACE_HINTS
                        )
                        cluster_info.likely_database_workloads = [
                            ns for ns in cluster_info.namespaces
                            if any(hint in ns.lower() for hint in _DB_NAMESPACE_HINTS)
                        ]
                    except Exception as e:
                        logging.warning(f"Could not fetch namespaces for {cluster.name}: {e}")

                    # ── PVCs ────────────────────────────────────────────────
                    try:
                        pvc_data = _kubectl(kubeconfig_file, "get", "pvc", "-A")
                        pvc_items = pvc_data["items"]
                        cluster_info.pvc_count = len(pvc_items)
                        cluster_info.pvc_names = [
                            f"{p['metadata'].get('namespace','')}/{p['metadata']['name']}"
                            for p in pvc_items
                        ]
                        cluster_info.pvc_unbound_count = sum(
                            1 for p in pvc_items if p.get("status", {}).get("phase") != "Bound"
                        )
                        pvc_size_total = 0.0
                        sc_map: dict = {}
                        for p in pvc_items:
                            sc = p["spec"].get("storageClassName", "unknown") or "unknown"
                            storage_str = (p.get("spec", {}).get("resources", {})
                                           .get("requests", {}).get("storage", "0Gi"))
                            size_gb = _parse_storage_gb(storage_str)
                            pvc_size_total += size_gb
                            if sc not in sc_map:
                                sc_map[sc] = {"count": 0, "size_gb": 0.0}
                            sc_map[sc]["count"] += 1
                            sc_map[sc]["size_gb"] = round(sc_map[sc]["size_gb"] + size_gb, 2)
                        cluster_info.total_pvc_size_gb = round(pvc_size_total, 2)
                        cluster_info.total_pvc_size_tb = round(pvc_size_total / 1024, 2)
                        cluster_info.storage_classes = list(sc_map.keys())
                        cluster_info.pvc_by_storage_class = sc_map
                    except Exception as e:
                        logging.warning(f"Could not fetch PVCs for {cluster.name}: {e}")

                    # ── StatefulSets ────────────────────────────────────────
                    try:
                        sts_data = _kubectl(kubeconfig_file, "get", "statefulsets", "-A")
                        cluster_info.statefulset_count = len(sts_data["items"])
                        cluster_info.statefulset_names = [
                            f"{s['metadata'].get('namespace','')}/{s['metadata']['name']}"
                            for s in sts_data["items"]
                        ]
                    except Exception as e:
                        logging.warning(f"Could not fetch StatefulSets for {cluster.name}: {e}")

                    # ── Deployments ─────────────────────────────────────────
                    try:
                        dep_data = _kubectl(kubeconfig_file, "get", "deployments", "-A")
                        cluster_info.deployment_count = len(dep_data["items"])
                    except Exception as e:
                        logging.warning(f"Could not fetch Deployments for {cluster.name}: {e}")

                    # ── ConfigMaps + Secrets (config complexity) ────────────
                    try:
                        cm_data = _kubectl(kubeconfig_file, "get", "configmaps", "-A")
                        cluster_info.configmap_count = len([
                            c for c in cm_data["items"]
                            if c["metadata"].get("namespace") not in {"kube-system", "kube-public"}
                        ])
                    except Exception as e:
                        logging.warning(f"Could not fetch ConfigMaps for {cluster.name}: {e}")
                    try:
                        sec_data = _kubectl(kubeconfig_file, "get", "secrets", "-A")
                        cluster_info.secret_count = len([
                            s for s in sec_data["items"]
                            if s["metadata"].get("namespace") not in {"kube-system", "kube-public"}
                            and s.get("type") != "kubernetes.io/service-account-token"
                        ])
                    except Exception as e:
                        logging.warning(f"Could not fetch Secrets for {cluster.name}: {e}")

                    # ── Backup Tool Detection ───────────────────────────────
                    try:
                        # Check for Velero/Kasten/Trilio by namespace or CRDs
                        all_ns_names = [n["metadata"]["name"] for n in (
                            _kubectl(kubeconfig_file, "get", "namespaces")["items"]
                        )]
                        detected = None
                        for tool, hints in _BACKUP_TOOL_LABELS.items():
                            if any(hint.split(".")[0] in ns for ns in all_ns_names for hint in hints):
                                detected = tool
                                break
                        # Also check CRDs for backup tool presence
                        if not detected:
                            try:
                                crd_data = _kubectl(kubeconfig_file, "get", "crds")
                                crd_names = [c["metadata"]["name"] for c in crd_data["items"]]
                                for tool, hints in _BACKUP_TOOL_LABELS.items():
                                    if any(any(h in crd for h in hints) for crd in crd_names):
                                        detected = tool
                                        break
                            except Exception:
                                pass
                        cluster_info.backup_tool_detected = detected or "none"
                        cluster_info.has_backup_annotations = detected is not None
                    except Exception as e:
                        logging.warning(f"Could not detect backup tools for {cluster.name}: {e}")
                        cluster_info.backup_tool_detected = "unknown"

                except Exception as e:
                    logging.warning(f"Error using kubectl for cluster {cluster.name}: {e}")
                finally:
                    try:
                        os.remove(kubeconfig_file)
                    except OSError:
                        pass

                region_summary.cluster_count += 1
                region_summary.total_node_count += cluster_info.node_count
                region_summary.total_pvc_count += cluster_info.pvc_count
                region_summary.total_pvc_size_gb += cluster_info.total_pvc_size_gb
                region_summary.total_pvc_size_tb += cluster_info.total_pvc_size_tb
                region_summary.total_statefulsets += cluster_info.statefulset_count
                if cluster_info.backup_tool_detected in ("none", "unknown"):
                    region_summary.unprotected_clusters += 1

                total_oke_clusters += 1
                total_oke_node_count += cluster_info.node_count
                total_oke_pvc_count += cluster_info.pvc_count
                total_oke_pvc_size_gb += cluster_info.total_pvc_size_gb
                total_oke_pvc_size_tb += cluster_info.total_pvc_size_tb
                total_oke_statefulsets += cluster_info.statefulset_count
                if cluster_info.backup_tool_detected in ("none", "unknown"):
                    total_oke_unprotected += 1

                compartment_oke_list.append(cluster_info)

            dump_info(filename, "oke_clusters", compartment_oke_list, include_sensitive)

        oke_summary_list.append(region_summary)

    write_grand_total(filename, "oke_clusters")
    dump_summary(filename, "oke_clusters", oke_summary_list)
    format_workbook(filename)
    logging.info("Completed processing OKE clusters.")
    logging.info(
        f"Grand Total - OKE Clusters: {total_oke_clusters}, "
        f"Nodes: {total_oke_node_count}, PVCs: {total_oke_pvc_count}, "
        f"PVC Size (GB): {total_oke_pvc_size_gb}, PVC Size (TB): {total_oke_pvc_size_tb}, "
        f"StatefulSets: {total_oke_statefulsets}, Unprotected: {total_oke_unprotected}"
    )


# ---------------------------------------------------------------
# AUTONOMOUS DATABASE
# ---------------------------------------------------------------
def get_autonomous_database_info(config, filename, regions=[], compartments=[]):
    identity_client = oci.identity.IdentityClient(config)
    if not regions:
        regions = [r.region_name for r in identity_client.list_region_subscriptions(config["tenancy"]).data]
    if not compartments:
        compartments = [c.id for c in identity_client.list_compartments(compartment_id=config["tenancy"], compartment_id_in_subtree=True).data]

    wb = load_workbook(filename)
    ws_info = wb.create_sheet("Autonomous DB Info") if "Autonomous DB Info" not in wb.sheetnames else wb["Autonomous DB Info"]
    ws_summary = wb.create_sheet("Autonomous DB Summary") if "Autonomous DB Summary" not in wb.sheetnames else wb["Autonomous DB Summary"]

    info_headers = ["Compartment ID", "ADB OCID", "Display Name", "Region", "DB Name", "Workload Type",
                    "Lifecycle State", "OCPU Count", "Storage (TB)", "Storage (GB)", "Is Free Tier",
                    "Is Dedicated", "DB Version", "Defined Tags", "Freeform Tags"]
    if ws_info.max_row == 1 and ws_info.cell(1, 1).value is None:
        ws_info.append(info_headers)
    if ws_summary.max_row == 1 and ws_summary.cell(1, 1).value is None:
        ws_summary.append(["Region", "ADB Count", "Total Storage (TB)", "Total Storage (GB)"])

    region_map: dict = {}
    for region in regions:
        config["region"] = region
        adb_client = oci.database.DatabaseClient(config)
        region_count, region_storage_tb = 0, 0.0
        for comp_id in compartments:
            try:
                adbs = oci.pagination.list_call_get_all_results(
                    adb_client.list_autonomous_databases, compartment_id=comp_id
                ).data
                for adb in adbs:
                    storage_tb = adb.data_storage_size_in_tbs or 0
                    storage_gb = round(storage_tb * 1024, 2)
                    ws_info.append([
                        comp_id, adb.id, adb.display_name, region, adb.db_name,
                        adb.db_workload, adb.lifecycle_state, adb.ocpu_count,
                        storage_tb, storage_gb, adb.is_free_tier, adb.is_dedicated,
                        adb.db_version, str(adb.defined_tags), str(adb.freeform_tags)
                    ])
                    region_count += 1
                    region_storage_tb += storage_tb
            except Exception as e:
                logging.warning(f"ADB error compartment {comp_id} region {region}: {e}")
        ws_summary.append([region, region_count, round(region_storage_tb, 4), round(region_storage_tb * 1024, 2)])
        region_map[region] = (region_count, region_storage_tb)

    total_adbs = sum(v[0] for v in region_map.values())
    total_tb = sum(v[1] for v in region_map.values())
    ws_summary.append(["TOTAL", total_adbs, round(total_tb, 4), round(total_tb * 1024, 2)])
    wb.save(filename)
    logging.info(f"Autonomous Database: {total_adbs} databases, {round(total_tb,4)} TB total")


# ---------------------------------------------------------------
# MYSQL HEATWAVE
# ---------------------------------------------------------------
def get_mysql_heatwave_info(config, filename, regions=[], compartments=[]):
    identity_client = oci.identity.IdentityClient(config)
    if not regions:
        regions = [r.region_name for r in identity_client.list_region_subscriptions(config["tenancy"]).data]
    if not compartments:
        compartments = [c.id for c in identity_client.list_compartments(compartment_id=config["tenancy"], compartment_id_in_subtree=True).data]

    wb = load_workbook(filename)
    ws_info = wb.create_sheet("MySQL HeatWave Info") if "MySQL HeatWave Info" not in wb.sheetnames else wb["MySQL HeatWave Info"]
    ws_summary = wb.create_sheet("MySQL HeatWave Summary") if "MySQL HeatWave Summary" not in wb.sheetnames else wb["MySQL HeatWave Summary"]

    if ws_info.max_row == 1 and ws_info.cell(1, 1).value is None:
        ws_info.append(["Compartment ID", "DB System ID", "Display Name", "Region", "Shape",
                        "Lifecycle State", "HA Mode", "HeatWave Enabled", "Node Count",
                        "Data Storage (GB)", "Defined Tags", "Freeform Tags"])
    if ws_summary.max_row == 1 and ws_summary.cell(1, 1).value is None:
        ws_summary.append(["Region", "DB System Count", "Total Storage (GB)"])

    region_map: dict = {}
    for region in regions:
        config["region"] = region
        mysql_client = oci.mysql.DbSystemClient(config)
        region_count, region_storage_gb = 0, 0.0
        for comp_id in compartments:
            try:
                dbs = oci.pagination.list_call_get_all_results(
                    mysql_client.list_db_systems, compartment_id=comp_id
                ).data
                for db in dbs:
                    storage_gb = db.data_storage_size_in_gbs or 0
                    hw_enabled = bool(db.heat_wave_cluster) if hasattr(db, 'heat_wave_cluster') else False
                    hw_nodes = db.heat_wave_cluster.cluster_size if hw_enabled and db.heat_wave_cluster else 0
                    ws_info.append([
                        comp_id, db.id, db.display_name, region, db.shape_name,
                        db.lifecycle_state, db.is_highly_available, hw_enabled, hw_nodes,
                        storage_gb, str(db.defined_tags), str(db.freeform_tags)
                    ])
                    region_count += 1
                    region_storage_gb += storage_gb
            except Exception as e:
                logging.warning(f"MySQL HeatWave error compartment {comp_id} region {region}: {e}")
        ws_summary.append([region, region_count, round(region_storage_gb, 2)])
        region_map[region] = (region_count, region_storage_gb)

    ws_summary.append(["TOTAL", sum(v[0] for v in region_map.values()), round(sum(v[1] for v in region_map.values()), 2)])
    wb.save(filename)
    logging.info(f"MySQL HeatWave: {sum(v[0] for v in region_map.values())} DB systems")


# ---------------------------------------------------------------
# NOSQL TABLES
# ---------------------------------------------------------------
def get_nosql_info(config, filename, regions=[], compartments=[]):
    identity_client = oci.identity.IdentityClient(config)
    if not regions:
        regions = [r.region_name for r in identity_client.list_region_subscriptions(config["tenancy"]).data]
    if not compartments:
        compartments = [c.id for c in identity_client.list_compartments(compartment_id=config["tenancy"], compartment_id_in_subtree=True).data]

    wb = load_workbook(filename)
    ws_info = wb.create_sheet("NoSQL Tables Info") if "NoSQL Tables Info" not in wb.sheetnames else wb["NoSQL Tables Info"]
    ws_summary = wb.create_sheet("NoSQL Tables Summary") if "NoSQL Tables Summary" not in wb.sheetnames else wb["NoSQL Tables Summary"]

    if ws_info.max_row == 1 and ws_info.cell(1, 1).value is None:
        ws_info.append(["Compartment ID", "Table ID", "Table Name", "Region", "Lifecycle State",
                        "Max Read Units", "Max Write Units", "Max Storage (GB)",
                        "Defined Tags", "Freeform Tags"])
    if ws_summary.max_row == 1 and ws_summary.cell(1, 1).value is None:
        ws_summary.append(["Region", "Table Count", "Total Max Storage (GB)"])

    region_map: dict = {}
    for region in regions:
        config["region"] = region
        nosql_client = oci.nosql.NosqlClient(config)
        region_count, region_storage_gb = 0, 0.0
        for comp_id in compartments:
            try:
                tables = oci.pagination.list_call_get_all_results(
                    nosql_client.list_tables, compartment_id=comp_id
                ).data
                for tbl in tables:
                    storage_gb = tbl.table_limits.max_storage_in_g_bs if tbl.table_limits else 0
                    ws_info.append([
                        comp_id, tbl.id, tbl.name, region, tbl.lifecycle_state,
                        tbl.table_limits.max_read_units if tbl.table_limits else 0,
                        tbl.table_limits.max_write_units if tbl.table_limits else 0,
                        storage_gb, str(tbl.defined_tags), str(tbl.freeform_tags)
                    ])
                    region_count += 1
                    region_storage_gb += storage_gb
            except Exception as e:
                logging.warning(f"NoSQL error compartment {comp_id} region {region}: {e}")
        ws_summary.append([region, region_count, round(region_storage_gb, 2)])
        region_map[region] = (region_count, region_storage_gb)

    ws_summary.append(["TOTAL", sum(v[0] for v in region_map.values()), round(sum(v[1] for v in region_map.values()), 2)])
    wb.save(filename)
    logging.info(f"NoSQL Tables: {sum(v[0] for v in region_map.values())} tables")


# ---------------------------------------------------------------
# FILE STORAGE SERVICE
# ---------------------------------------------------------------
def get_file_storage_info(config, filename, regions=[], compartments=[]):
    identity_client = oci.identity.IdentityClient(config)
    if not regions:
        regions = [r.region_name for r in identity_client.list_region_subscriptions(config["tenancy"]).data]
    if not compartments:
        compartments = [c.id for c in identity_client.list_compartments(compartment_id=config["tenancy"], compartment_id_in_subtree=True).data]

    wb = load_workbook(filename)
    ws_info = wb.create_sheet("File Storage Info") if "File Storage Info" not in wb.sheetnames else wb["File Storage Info"]
    ws_summary = wb.create_sheet("File Storage Summary") if "File Storage Summary" not in wb.sheetnames else wb["File Storage Summary"]

    if ws_info.max_row == 1 and ws_info.cell(1, 1).value is None:
        ws_info.append(["Compartment ID", "File System ID", "Display Name", "Region",
                        "Availability Domain", "Lifecycle State",
                        "Metered Bytes", "Size (GB)", "Size (TB)",
                        "Defined Tags", "Freeform Tags"])
    if ws_summary.max_row == 1 and ws_summary.cell(1, 1).value is None:
        ws_summary.append(["Region", "File System Count", "Total Size (GB)", "Total Size (TB)"])

    region_map: dict = {}
    for region in regions:
        config["region"] = region
        fss_client = oci.file_storage.FileStorageClient(config)
        region_count, region_size_gb = 0, 0.0

        # File Storage is AZ-scoped — must iterate ADs
        try:
            ads = identity_client.list_availability_domains(config["tenancy"]).data
        except Exception:
            ads = []

        for comp_id in compartments:
            for ad in ads:
                try:
                    fss_list = oci.pagination.list_call_get_all_results(
                        fss_client.list_file_systems,
                        compartment_id=comp_id,
                        availability_domain=ad.name
                    ).data
                    for fs in fss_list:
                        metered_bytes = fs.metered_bytes or 0
                        size_gb = round(metered_bytes / 1e9, 4)
                        size_tb = round(metered_bytes / 1e12, 4)
                        ws_info.append([
                            comp_id, fs.id, fs.display_name, region,
                            ad.name, fs.lifecycle_state,
                            metered_bytes, size_gb, size_tb,
                            str(fs.defined_tags), str(fs.freeform_tags)
                        ])
                        region_count += 1
                        region_size_gb += size_gb
                except Exception as e:
                    logging.warning(f"File Storage error AD {ad.name} compartment {comp_id} region {region}: {e}")

        ws_summary.append([region, region_count, round(region_size_gb, 4), round(region_size_gb / 1024, 4)])
        region_map[region] = (region_count, region_size_gb)

    total_count = sum(v[0] for v in region_map.values())
    total_gb = sum(v[1] for v in region_map.values())
    ws_summary.append(["TOTAL", total_count, round(total_gb, 4), round(total_gb / 1024, 4)])
    wb.save(filename)
    logging.info(f"File Storage: {total_count} file systems, {round(total_gb,2)} GB total")


if __name__ == "__main__":

    if not shutil.which("kubectl"):
        logging.error("Error: 'kubectl' command not found. Please install kubectl to proceed.")
        sys.exit(1)
    if not shutil.which("oci"):
        logging.error("Error: 'oci' CLI not found. Please install OCI CLI to proceed.")
        sys.exit(1)

    packages = ["oci", "openpyxl", "pandas"]

    for pkg in packages:
        install_and_import(pkg)

    args = sys.argv[1:]
    for arg in args:
        if arg.startswith("--profile="):
            profile_name = arg.split("=")[1]
        elif arg.startswith("--region="):
            regions = arg.split("=")[1].split(",")
        elif arg.startswith("--compartment="):
            compartments = arg.split("=")[1].split(",")
        elif arg.startswith("--workload="):
            workload = arg.split("=")[1]
        elif arg.startswith("--output-format="):
            output_format = arg.split("=")[1].lower()
            if output_format not in ("csv", "json", "both"):
                print(f"Invalid --output-format '{output_format}'. Valid values: csv, json, both")
                sys.exit(1)
        elif arg == "--include-sensitive":
            include_sensitive = True
        elif arg == "--help":
            print(
                "Usage: python CVOracleCloudSizingScript.py "
                "[--workload=<instances|object_storage|db_systems|oke_clusters|autonomous_db|mysql_heatwave|nosql|file_storage|all>] "
                "[--profile=<profilename>] [--region=<region1>,<region2>] [--compartment=<comp1>,<comp2>] "
                "[--output-format=<csv|json|both>] [--include-sensitive] [--help]\n"
                "\n"
                "  --include-sensitive  Include potentially private data: namespace names, PVC names,\n"
                "                       node names, and StatefulSet names in the report.\n"
                "                       Omitted by default for data privacy."
            )
            sys.exit(0)
        else:
            print(f"Unknown argument: {arg}")
            sys.exit(1)

    if "workload" not in locals():
        workload = "all"
    if "profile_name" not in locals():
        profile_name = oci.config.DEFAULT_PROFILE
    if "regions" not in locals():
        regions = []
    if "compartments" not in locals():
        compartments = []
    if "output_format" not in locals():
        output_format = "csv"
    if "include_sensitive" not in locals():
        include_sensitive = False

    if not include_sensitive:
        logging.info(
            "Sensitive fields (namespace names, PVC names, node names) are HIDDEN. "
            "Pass --include-sensitive to include them."
        )

    config = oci.config.from_file(profile_name=profile_name)
    timestamp = datetime.now().strftime("%Y%m%d%H%M%S")
    log_dir = "Logs"
    os.makedirs(log_dir, exist_ok=True)
    log_filename = os.path.join(log_dir, f"{profile_name}_{workload}_{timestamp}.log")
    handlers=[
        logging.FileHandler(log_filename),
        logging.StreamHandler(sys.stdout)
    ]
    logging.basicConfig(
        handlers=handlers,
        level=logging.INFO,
        format='%(asctime)s - %(levelname)s - %(message)s'
    )
    metrics_dir = "Metrics"
    os.makedirs(metrics_dir, exist_ok=True)
    filename = os.path.join(metrics_dir, f"{profile_name}_{workload}_{timestamp}.xlsx")

    # Track workload summaries for JSON export
    json_summary: dict = {}
    json_workloads: dict = {}

    def _tracked_runner(wl_key, runner_fn):
        """Wrap a workload runner to capture summary stats for JSON export."""
        def _inner():
            runner_fn()
            if output_format in ("json", "both"):
                try:
                    wb = load_workbook(filename)
                    # Read summary sheet rows if available
                    summary_sheet_name = None
                    for sn in wb.sheetnames:
                        if "summary" in sn.lower() and wl_key.split("_")[0].lower() in sn.lower():
                            summary_sheet_name = sn
                            break
                    count, total_gb = 0, 0.0
                    if summary_sheet_name:
                        ws = wb[summary_sheet_name]
                        for row in ws.iter_rows(min_row=2, values_only=True):
                            if row and str(row[0] or "").upper() == "TOTAL":
                                count = int(row[1] or 0) if len(row) > 1 else 0
                                total_gb = float(row[3] or 0) if len(row) > 3 else float(row[2] or 0)
                                break
                    json_summary[wl_key] = {"count": count, "total_storage_gb": round(total_gb, 2), "notes": ""}
                except Exception as ex:
                    logging.warning(f"JSON summary capture failed for {wl_key}: {ex}")
        return _inner

    WORKLOAD_RUNNERS = {
        "instances":       lambda: (init_excel(filename, "instances"),       get_instance_info(config, filename, regions, compartments)),
        "object_storage":  lambda: (init_excel(filename, "object_storage"),  get_object_storage_info(config, filename, regions, compartments)),
        "db_systems":      lambda: (init_excel(filename, "db_systems"),       get_database_info(config, filename, regions, compartments)),
        "oke_clusters":    lambda: (init_excel(filename, "oke_clusters"),     get_oke_cluster_info(config, filename, regions, compartments)),
        "autonomous_db":   lambda: get_autonomous_database_info(config, filename, regions, compartments),
        "mysql_heatwave":  lambda: get_mysql_heatwave_info(config, filename, regions, compartments),
        "nosql":           lambda: get_nosql_info(config, filename, regions, compartments),
        "file_storage":    lambda: get_file_storage_info(config, filename, regions, compartments),
    }

    if workload == "all":
        logging.info("Getting information for all supported workloads.")
        for wl, runner in WORKLOAD_RUNNERS.items():
            logging.info(f"--- Starting workload: {wl} ---")
            try:
                runner()
            except Exception as e:
                logging.error(f"Workload {wl} failed: {e}")
    elif workload in WORKLOAD_RUNNERS:
        WORKLOAD_RUNNERS[workload]()
    else:
        supported = ", ".join(WORKLOAD_RUNNERS.keys()) + ", all"
        logging.error(f"Unsupported workload: {workload}. Supported: {supported}")
        sys.exit(1)

    # ----------------------------------------------------------------
    # JSON EXPORT (when --output-format=json or both)
    # ----------------------------------------------------------------
    if output_format in ("json", "both"):
        try:
            # Read summary stats from the Excel workbook sheets
            wb = load_workbook(filename)
            workload_sheet_map = {
                "instances":      ("Instance Summary",),
                "object_storage": ("Object Storage Summary",),
                "db_systems":     ("DB System Summary",),
                "oke_clusters":   ("OKE Cluster Summary",),
                "autonomous_db":  ("Autonomous DB Summary",),
                "mysql_heatwave": ("MySQL HeatWave Summary",),
                "nosql":          ("NoSQL Tables Summary",),
                "file_storage":   ("File Storage Summary",),
            }

            # Column indices for the total row in each summary sheet:
            # (count_col, storage_gb_col) — 0-based, matching the summary_headers order
            wl_col_map = {
                "instances":      (1, 2),   # Instance Count, Total Size (GB)
                "object_storage": (1, 2),   # Bucket Count, Total Storage (GB)
                "db_systems":     (1, 2),   # DB System Count, Total Storage (GB)
                "oke_clusters":   (1, 4),   # Cluster Count (col1), Total PVC Size GB (col4)
                "autonomous_db":  (1, 2),
                "mysql_heatwave": (1, 2),
                "nosql":          (1, 2),
                "file_storage":   (1, 2),
            }
            for wl_key, sheet_names in workload_sheet_map.items():
                for sheet_name in sheet_names:
                    if sheet_name in wb.sheetnames:
                        ws = wb[sheet_name]
                        count, total_gb = 0, 0.0
                        count_col, gb_col = wl_col_map.get(wl_key, (1, 3))
                        for row in ws.iter_rows(min_row=2, values_only=True):
                            if not row or row[0] is None:
                                continue
                            label = str(row[0]).upper()
                            # Match any "TOTAL" row (e.g. "Total OKE Clusters", "TOTAL", etc.)
                            if "TOTAL" in label:
                                count = int(row[count_col] or 0) if len(row) > count_col else 0
                                total_gb = float(row[gb_col] or 0) if len(row) > gb_col else 0.0
                                break
                        json_summary[wl_key] = {"count": count, "total_storage_gb": round(total_gb, 2), "notes": ""}

            # ── Build rich per-cluster workload detail from Excel ───────
            _SENSITIVE_COLUMNS = {"Namespaces", "StatefulSets", "PVC Names", "Node Names"}
            json_workloads: dict = {}
            if "OKE Cluster Info" in wb.sheetnames:
                ws_info = wb["OKE Cluster Info"]
                headers = [c.value for c in next(ws_info.iter_rows(min_row=1, max_row=1))]
                clusters_detail = []
                for row in ws_info.iter_rows(min_row=2, values_only=True):
                    if not row or row[0] is None:
                        continue
                    entry = {}
                    for col_name, val in zip(headers, row):
                        if col_name in _SENSITIVE_COLUMNS and not include_sensitive:
                            entry[col_name] = "[redacted]"
                        else:
                            entry[col_name] = val
                    clusters_detail.append(entry)
                json_workloads["oke_clusters"] = clusters_detail

            # ── Protection gap analysis ─────────────────────────────────
            protection_gaps = []
            for cluster in json_workloads.get("oke_clusters", []):
                gaps = []
                tool = str(cluster.get("Backup Tool Detected") or "none").lower()
                sts = cluster.get("StatefulSet Count") or 0
                pvc_gb = cluster.get("Total PVC Size (GB)") or 0
                unbound = cluster.get("Unbound PVCs") or 0
                is_prod = str(cluster.get("Likely Production") or "").lower() == "true"
                db_ns = cluster.get("Likely DB Workload Namespaces") or ""

                if tool in ("none", "unknown"):
                    gaps.append("NO_BACKUP_TOOL: No Velero/Kasten/Trilio detected — cluster has no automated backup")
                if int(sts) > 0 and tool in ("none", "unknown"):
                    gaps.append(f"STATEFULSET_UNPROTECTED: {sts} StatefulSets with no backup tool")
                if int(unbound) > 0:
                    gaps.append(f"UNBOUND_PVCS: {unbound} PVCs not bound — potential data loss risk")
                if is_prod and tool in ("none", "unknown"):
                    gaps.append("PRODUCTION_UNPROTECTED: Production namespaces detected with no backup coverage")
                if db_ns:
                    gaps.append(f"DATABASE_WORKLOADS: DB-like namespaces without dedicated backup: {db_ns}")

                protection_gaps.append({
                    "cluster": cluster.get("Cluster Name"),
                    "region": cluster.get("Region"),
                    "backup_tool": tool,
                    "statefulsets": sts,
                    "pvc_count": cluster.get("PVC Count") or 0,
                    "pvc_size_gb": pvc_gb,
                    "is_production": is_prod,
                    "gaps": gaps,
                    "risk_score": len(gaps),  # 0 = protected, higher = more exposure
                })

            # Sort by risk score descending
            protection_gaps.sort(key=lambda x: x["risk_score"], reverse=True)

            json_doc = {
                "metadata": {
                    "cloud": "oci",
                    "tenancy": config.get("tenancy", ""),
                    "regions": regions if regions else ["all"],
                    "generated_at": datetime.now().isoformat(),
                    "script_version": "3.0",
                    "workload": workload,
                    "sensitive_fields_included": include_sensitive,
                },
                "summary": json_summary,
                "workloads": json_workloads,
                "protection_gap_analysis": protection_gaps,
                "excel_report": filename,
            }

            json_path = os.path.join(metrics_dir, f"oci_sizing_{timestamp}.json")
            with open(json_path, "w", encoding="utf-8") as f:
                json.dump(json_doc, f, indent=2, default=str)
            logging.info(f"JSON sizing report written: {json_path}")
        except Exception as ex:
            logging.error(f"JSON export failed: {ex}")